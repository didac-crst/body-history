from __future__ import annotations

import hashlib
import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .metrics import convert_fraction_to_percent
from .models import ImportBatch, Measurement, MeasurementImportRow, Profile

logger = logging.getLogger(__name__)

SOURCE_SHEET = "General"
COLUMN_MAP = {
    "Data": "measured_at",
    "Pes": "weight_kg",
    "% Grassa": "body_fat_percent",
    "% Muscul": "muscle_percent",
    "Alçada": "height_m",
    "Outlier?": "outlier",
}
IGNORED_FORMULA_COLUMNS = {
    "Grassa",
    "IMC",
    "Nota Pes",
    "Nota Grassa",
    "Nota Muscul",
    "Nota Total",
    "Year",
}


@dataclass
class ParsedRow:
    source_row: int
    measured_at: datetime | None = None
    weight_kg: Decimal | None = None
    body_fat_percent: Decimal | None = None
    muscle_percent: Decimal | None = None
    height_cm: Decimal | None = None
    outlier: Any = None
    raw_payload: dict = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors and self.measured_at is not None and self.weight_kg is not None


@dataclass
class DryRunReport:
    sheets: list[str]
    source_sheet: str
    file_hash: str
    filename: str
    column_mappings: dict
    formulas_ignored: list[str]
    candidate_count: int
    accepted_count: int
    rejected_count: int
    date_range: dict
    missing_values: dict
    invalid_values: list[dict]
    duplicate_local_dates: list[str]
    weight_range: dict
    height_cm: Decimal | None
    expected_hints: dict
    rows: list[ParsedRow] = field(default_factory=list)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise ValueError(f"not a number: {value!r}")


def _as_datetime(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if timezone.is_naive(value):
            return timezone.make_aware(value, timezone.get_current_timezone())
        return value
    if isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
        return timezone.make_aware(dt, timezone.get_current_timezone())
    raise ValueError(f"not a date: {value!r}")


def parse_workbook(path: Path) -> DryRunReport:
    path = Path(path)
    digest = file_sha256(path)
    wb_formulas = load_workbook(path, data_only=False, read_only=True)
    sheets = list(wb_formulas.sheetnames)
    if SOURCE_SHEET not in sheets:
        wb_formulas.close()
        raise ValueError(f"Expected sheet {SOURCE_SHEET!r} not found in {sheets}")

    formula_sheet = wb_formulas[SOURCE_SHEET]
    header_cells = next(formula_sheet.iter_rows(min_row=1, max_row=1, values_only=False))
    headers = [cell.value for cell in header_cells]
    formula_flags = []
    for row in formula_sheet.iter_rows(min_row=2, max_row=3, values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                col_name = headers[cell.column - 1] if cell.column <= len(headers) else None
                if col_name and col_name not in formula_flags:
                    formula_flags.append(col_name)
    wb_formulas.close()

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[SOURCE_SHEET]
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: i for i, name in enumerate(header) if name}

    rows: list[ParsedRow] = []
    local_dates: dict[str, list[int]] = {}
    missing = {"body_fat_percent": 0, "muscle_percent": 0, "weight_kg": 0, "measured_at": 0}
    invalid: list[dict] = []
    weights: list[Decimal] = []
    heights: list[Decimal] = []

    for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if values is None or all(v is None or v == "" for v in values[:7]):
            continue
        raw = {
            str(header[i]): _jsonable(values[i]) if i < len(values) else None
            for i in range(min(len(header), 13))
            if header[i]
        }
        parsed = ParsedRow(source_row=excel_row, raw_payload=raw)

        try:
            parsed.measured_at = _as_datetime(values[index["Data"]] if "Data" in index else None)
        except (ValueError, KeyError) as exc:
            parsed.errors.append(f"measured_at: {exc}")

        try:
            parsed.weight_kg = _to_decimal(values[index["Pes"]] if "Pes" in index else None)
            if parsed.weight_kg is not None and parsed.weight_kg <= 0:
                parsed.errors.append("weight_kg must be > 0")
        except (ValueError, KeyError) as exc:
            parsed.errors.append(f"weight_kg: {exc}")

        try:
            raw_fat = values[index["% Grassa"]] if "% Grassa" in index else None
            if raw_fat in (0, "0", 0.0):
                raw_fat = None
            parsed.body_fat_percent = convert_fraction_to_percent(raw_fat)
        except (ValueError, KeyError, InvalidOperation) as exc:
            parsed.errors.append(f"body_fat_percent: {exc}")

        try:
            raw_muscle = values[index["% Muscul"]] if "% Muscul" in index else None
            if raw_muscle in (0, "0", 0.0):
                raw_muscle = None
            parsed.muscle_percent = convert_fraction_to_percent(raw_muscle)
        except (ValueError, KeyError, InvalidOperation) as exc:
            parsed.errors.append(f"muscle_percent: {exc}")

        try:
            height_m = _to_decimal(values[index["Alçada"]] if "Alçada" in index else None)
            if height_m is not None:
                # Workbook stores metres (1.81); profile stores cm.
                parsed.height_cm = (height_m * Decimal("100")).quantize(Decimal("0.01"))
                heights.append(parsed.height_cm)
        except (ValueError, KeyError) as exc:
            parsed.warnings.append(f"height: {exc}")

        if "Outlier?" in index:
            parsed.outlier = values[index["Outlier?"]]

        if parsed.measured_at is None:
            missing["measured_at"] += 1
            parsed.errors.append("missing measured_at")
        if parsed.weight_kg is None:
            missing["weight_kg"] += 1
            parsed.errors.append("missing weight_kg")
        if parsed.body_fat_percent is None:
            missing["body_fat_percent"] += 1
        if parsed.muscle_percent is None:
            missing["muscle_percent"] += 1

        if parsed.measured_at is not None:
            key = timezone.localtime(parsed.measured_at).date().isoformat()
            local_dates.setdefault(key, []).append(excel_row)

        if parsed.weight_kg is not None:
            weights.append(parsed.weight_kg)

        if not parsed.ok:
            invalid.append(
                {
                    "source_row": excel_row,
                    "errors": parsed.errors,
                }
            )

        rows.append(parsed)

    wb.close()

    accepted = [r for r in rows if r.ok]
    rejected = [r for r in rows if not r.ok]
    dates = [
        timezone.localtime(r.measured_at).date()
        for r in accepted
        if r.measured_at is not None
    ]
    dups = sorted(date for date, idxs in local_dates.items() if len(idxs) > 1)

    height_cm = None
    if heights:
        # Prefer the modal/latest non-null height.
        height_cm = heights[-1]

    report = DryRunReport(
        sheets=sheets,
        source_sheet=SOURCE_SHEET,
        file_hash=digest,
        filename=path.name,
        column_mappings=COLUMN_MAP,
        formulas_ignored=sorted(set(formula_flags) | IGNORED_FORMULA_COLUMNS),
        candidate_count=len(rows),
        accepted_count=len(accepted),
        rejected_count=len(rejected),
        date_range={
            "first": dates[0].isoformat() if dates else None,
            "last": dates[-1].isoformat() if dates else None,
            "unique_dates": len(set(dates)),
        },
        missing_values=missing,
        invalid_values=invalid,
        duplicate_local_dates=dups,
        weight_range={
            "min": float(min(weights)) if weights else None,
            "max": float(max(weights)) if weights else None,
        },
        height_cm=height_cm,
        expected_hints={
            "unique_measurement_dates": 691,
            "first_date": "2005-04-15",
            "last_date": "2025-11-15",
            "weight_min": 70.0,
            "weight_max": 82.8,
        },
        rows=rows,
    )
    return report


def _jsonable(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def report_to_dict(report: DryRunReport) -> dict:
    return {
        "sheets": report.sheets,
        "source_sheet": report.source_sheet,
        "file_hash": report.file_hash,
        "filename": report.filename,
        "column_mappings": report.column_mappings,
        "formulas_ignored": report.formulas_ignored,
        "candidate_count": report.candidate_count,
        "accepted_count": report.accepted_count,
        "rejected_count": report.rejected_count,
        "date_range": report.date_range,
        "missing_values": report.missing_values,
        "invalid_values": report.invalid_values,
        "duplicate_local_dates": report.duplicate_local_dates,
        "weight_range": report.weight_range,
        "height_cm": float(report.height_cm) if report.height_cm is not None else None,
        "expected_hints": report.expected_hints,
        "matches_expected_unique_dates": report.date_range.get("unique_dates")
        == report.expected_hints["unique_measurement_dates"],
    }


@transaction.atomic
def import_workbook(profile: Profile, path: Path, report: DryRunReport | None = None) -> ImportBatch:
    path = Path(path)
    report = report or parse_workbook(path)

    existing = ImportBatch.objects.filter(profile=profile, file_hash=report.file_hash).first()
    if existing:
        logger.info("Import skipped; file hash already imported for profile")
        return existing

    if report.height_cm is not None:
        profile.height_cm = report.height_cm
        profile.save(update_fields=["height_cm", "updated_at"])

    batch = ImportBatch.objects.create(
        profile=profile,
        filename=report.filename,
        file_hash=report.file_hash,
        row_count=report.candidate_count,
        accepted_count=0,
        rejected_count=0,
        metadata=report_to_dict(report),
    )

    accepted = 0
    rejected = 0
    for parsed in report.rows:
        if not parsed.ok:
            MeasurementImportRow.objects.create(
                import_batch=batch,
                source_sheet=SOURCE_SHEET,
                source_row=parsed.source_row,
                raw_payload=parsed.raw_payload,
                status=MeasurementImportRow.STATUS_REJECTED,
                error_message="; ".join(parsed.errors),
            )
            rejected += 1
            continue

        legacy = {
            "outlier": parsed.outlier,
            "source_row": parsed.source_row,
            "raw": parsed.raw_payload,
        }
        measurement = Measurement.objects.create(
            profile=profile,
            measured_at=parsed.measured_at,
            weight_kg=parsed.weight_kg,
            body_fat_percent=parsed.body_fat_percent,
            muscle_percent=parsed.muscle_percent,
            source=Measurement.SOURCE_IMPORT,
            legacy_payload=legacy,
        )
        MeasurementImportRow.objects.create(
            import_batch=batch,
            source_sheet=SOURCE_SHEET,
            source_row=parsed.source_row,
            raw_payload=parsed.raw_payload,
            measurement=measurement,
            status=MeasurementImportRow.STATUS_ACCEPTED,
        )
        accepted += 1

    batch.accepted_count = accepted
    batch.rejected_count = rejected
    batch.save(update_fields=["accepted_count", "rejected_count"])
    return batch
